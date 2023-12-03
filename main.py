import requests
import json
import pandas as pd
from tkinter import Tk, StringVar, OptionMenu, Button
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class JiraProjectSelector:
    def __init__(self, url, username, password):
        self.url = url
        self.username = username
        self.password = password
        self.root = Tk()
        self.root.configure(background="#E8E8E8")
        self.root.title("Выбор проекта Jira")
        self.root.geometry("300x150")

    def get_projects(self):
        response = requests.get(f"{self.url}/rest/api/2/project", auth=(self.username, self.password))
        projects = response.json()
        return projects

    def show_project_selector(self):
        projects = self.get_projects()
        project_names = [project['name'] for project in projects]
        project_keys = [project['key'] for project in projects]

        selected_project = StringVar(self.root)
        selected_project.set(project_keys[0])

        option_menu = OptionMenu(self.root, selected_project, *project_names)
        option_menu.config(bg="#FFFFFF")
        option_menu.pack(pady=20)

        submit_button = Button(self.root, text="Submit", command=self.submit)
        submit_button.config(bg="#4CAF50", fg="#FFFFFF")
        submit_button.pack()

        self.root.mainloop()

        for i, name in enumerate(project_names):
            if selected_project.get() == name:
                return project_keys[i]

    def submit(self):
        self.root.destroy()


# Пример использования
url = "https://jira.fsk.ru"
username = 'belopakhovas'
password = 'Silva2001'

project_selector = JiraProjectSelector(url, username, password)
selected_project = project_selector.show_project_selector()

project = selected_project


class JiraWorklogData:
    def __init__(self, url, username, password):
        self.url = url
        self.username = username
        self.password = password
        self.data = None
        self.time_spent_dict = {}

    def fetch_data(self):
        response = requests.get(self.url, auth=(self.username, self.password))
        self.data = response.json()

    def process_data(self):
        for issue in self.data['issues']:
            for log in issue['fields']['worklog']['worklogs']:
                performer_name = log['author']['displayName']
                time_spent = log['timeSpent']

                if performer_name in self.time_spent_dict:
                    self.time_spent_dict[performer_name] += pd.to_timedelta(time_spent)
                else:
                    self.time_spent_dict[performer_name] = pd.to_timedelta(time_spent)

    def extract_results(self):
        task_code = []
        performer_name_list = []
        performer_ids = []
        total_time_spent = []

        for issue in self.data['issues']:
            unique_performers = set()
            for log in issue['fields']['worklog']['worklogs']:
                performer_name = log['author']['displayName']
                performer_id = log['author']['key']

                if performer_name in unique_performers:
                    continue

                unique_performers.add(performer_name)

                task_code.append(issue['key'])
                performer_name_list.append(performer_name)
                performer_ids.append(performer_id)
                total_time_spent.append(str(self.time_spent_dict[performer_name]))

        df = pd.DataFrame({'Task Code': task_code, 'Performer Name': performer_name_list, 'Performer ID': performer_ids,
                           'Total Time Spent': total_time_spent})
        df = df.drop_duplicates(subset='Performer Name')
        return df

    def save_to_excel(self, filename):
        df = self.extract_results()
        df.to_excel(filename, index=False)
        print(f"Результаты сохранены в файл: {filename}")

# Использование класса JiraWorklogData
url = f'https://jira.fsk.ru/rest/api/2/search?jql=project=%22{project}%22%20AND%20worklogDate%20%3E=%20startOfMonth()%20AND%20worklogDate%20%3C=%20endOfMonth()&fields=worklog&maxResults=1000'
username = 'belopakhovas'
password = 'Silva2001'
filename = 'results9.xlsx'

worklog_data = JiraWorklogData(url, username, password)
worklog_data.fetch_data()
worklog_data.process_data()
worklog_data.save_to_excel(filename)

class ExcelProcessor:
    def __init__(self, file_path, threshold_hours):
        self.file_path = file_path
        self.threshold_hours = threshold_hours
        self.df = None
        self.wb = None
        self.ws = None

    def read_excel(self):
        self.df = pd.read_excel(self.file_path)

    def convert_time_to_hours(self):
        self.df['Время (ч)'] = pd.to_timedelta(self.df['Total Time Spent']).dt.total_seconds() / 3600

    def highlight_rows(self):
        red_fill = PatternFill(fill_type='solid', fgColor='FF0000')

        for index, row in self.df.iterrows():
            time_hours = row['Время (ч)']
            for col_num, value in enumerate(row, start=1):
                self.ws.cell(row=index + 2, column=col_num).value = value
                if time_hours < self.threshold_hours:
                    self.ws.cell(row=index + 2, column=col_num).fill = red_fill

    def save_excel(self, new_file_path):
        self.wb.save(new_file_path)

    def process_excel(self):
        self.read_excel()
        self.convert_time_to_hours()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.highlight_rows()

file_path = 'results9.xlsx'
threshold_hours = 15

excel_processor = ExcelProcessor(file_path, threshold_hours)
excel_processor.process_excel()
new_file_path = 'Итог.xlsx'
excel_processor.save_excel(new_file_path)